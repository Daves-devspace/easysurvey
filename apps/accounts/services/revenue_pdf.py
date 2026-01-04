# apps/EasyDocs/pdf_utils.py
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from datetime import datetime  # ✅ Correct import for strptime
from django.utils import timezone
from apps.EasyDocs.accounts.revenue import get_revenue_from_payments
from decimal import Decimal
from django.http import HttpResponse
import tempfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_revenue_pdf(buffer, start_date, end_date, status_filter='all'):
    """
    Generate a clean, landscape PDF listing all revenue records.
    Optimized for data fitting and pagination.
    """
    # 1. Fetch data
    revenue_data = get_revenue_from_payments(
        start_date=start_date,
        end_date=end_date,
        profit_mode="auto"
    )

    # 2. ✅ Apply Status Filter
    qs = revenue_data.get('revenue_qs')
    if qs is not None and status_filter == 'completed':
        qs = qs.filter(status='completed')
        revenue_data['revenue_qs'] = qs

    # 3. Setup Document (Landscape for wider tables)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elements = []
    styles = getSampleStyleSheet()
    
    # Custom Styles
    style_title = styles['Title']
    
    # Style for table text
    style_cell = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        alignment=TA_LEFT
    )
    
    # Report Header
    status_label = " (Completed Only)" if status_filter == 'completed' else ""
    title_text = f"Revenue Report{status_label}: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    elements.append(Paragraph(title_text, style_title))
    elements.append(Spacer(1, 20))

    # Summary Table (Totals)
    total_col_widths = [150, 120, 120, 120]
    total_data = [
        ['Category', 'Gross (KES)', 'Company (KES)', 'Institution (KES)'],
        [
            'Main Services',
            f"{revenue_data['main_services']['gross_total']:.2f}",
            f"{revenue_data['main_services']['company_total']:.2f}",
            f"{revenue_data['main_services']['inst_total']:.2f}"
        ],
        [
            'Sub Services',
            f"{revenue_data['subservices']['gross_total']:.2f}",
            f"{revenue_data['subservices']['company_total']:.2f}",
            f"{revenue_data['subservices']['inst_total']:.2f}"
        ],
        [
            'TOTAL',
            f"{revenue_data['gross_total']:.2f}",
            f"{revenue_data['company_total']:.2f}",
            f"{revenue_data['inst_total']:.2f}"
        ]
    ]
    
    t_summary = Table(total_data, colWidths=total_col_widths, hAlign='CENTER')
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.whitesmoke),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t_summary)
    elements.append(Spacer(1, 30))

    # Detailed Records
    elements.append(Paragraph("Detailed Revenue Records", styles['Heading2']))
    elements.append(Spacer(1, 10))

    col_widths = [30, 160, 190, 100, 100, 100, 60]
    headers = ['#', 'Client', 'Service', 'Client Charge', 'Inst. Cost', 'Profit', '%']
    table_data = [headers]
    
    records = revenue_data.get('revenue_qs', [])
    
    if not records:
        table_data.append(['-', 'No records found', '', '', '', '', ''])
    else:
        for idx, record in enumerate(records, 1):
            client_p = Paragraph(f"{record.client.first_name} {record.client.last_name}", style_cell)
            service_p = Paragraph(record.service.name, style_cell)
            
            charge = f"{record.client_charge:,.2f}"
            inst = f"{record.inst_cost:,.2f}"
            profit = f"{record.profit_amount:,.2f}"
            margin = f"{record.profit_percent:.1f}%"
            
            row = [str(idx), client_p, service_p, charge, inst, profit, margin]
            table_data.append(row)

    t_records = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign='LEFT')
    t_records.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    elements.append(t_records)
    doc.build(elements)
    return buffer

def revenue_pdf_view(request):
    """Generate PDF for the current filters"""
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    status_filter = request.GET.get('status', 'all')
    
    try:
        if not start_date_str or not end_date_str:
            raise ValueError("Missing dates")
            
        # 1. Try ISO format first (YYYY-MM-DD)
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            # 2. Fallback: Try verbose format like "Jan. 1, 2025" or "Jan 1, 2025"
            # Removing dots allows 'Jan.' to match '%b' (which usually expects 'Jan')
            clean_start = start_date_str.replace('.', '')
            clean_end = end_date_str.replace('.', '')
            start_date = datetime.strptime(clean_start, "%b %d, %Y").date()
            end_date = datetime.strptime(clean_end, "%b %d, %Y").date()
            
    except (ValueError, TypeError):
        # Fallback to current year only if parsing actually fails
        start_date = timezone.now().date().replace(month=1, day=1)
        end_date = timezone.now().date()

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf_path = generate_revenue_pdf(tmp_file.name, start_date, end_date, status_filter)

    with open(pdf_path, 'rb') as f:
        pdf_data = f.read()

    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="revenue_{start_date}_{end_date}.pdf"'
    return response

def generate_revenue_excel(start_date, end_date, status_filter='all'):
    """Generate a clean, well-formatted Excel revenue report."""
    revenue_data = get_revenue_from_payments(
        start_date=start_date,
        end_date=end_date,
        profit_mode="auto"
    )

    qs = revenue_data.get('revenue_qs')
    if qs is not None and status_filter == 'completed':
        qs = qs.filter(status='completed')
        revenue_data['revenue_qs'] = qs

    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    status_label = " (Completed Only)" if status_filter == 'completed' else ""
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Revenue Report{status_label}: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])

    # Totals
    ws.append(["Category", "Gross Revenue (KES)", "Company Revenue (KES)", "Institution Revenue (KES)"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    totals = [
        ["Main Services", revenue_data['main_services']['gross_total'], revenue_data['main_services']['company_total'], revenue_data['main_services']['inst_total']],
        ["Sub Services", revenue_data['subservices']['gross_total'], revenue_data['subservices']['company_total'], revenue_data['subservices']['inst_total']],
        ["TOTAL", revenue_data['gross_total'], revenue_data['company_total'], revenue_data['inst_total']]
    ]

    for row in totals:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            if row[0] == "TOTAL":
                cell.font = Font(bold=True)
                cell.fill = total_fill
    ws.append([])

    # Detailed Records
    ws.append(["#", "Client", "Service", "Client Charge", "Inst. Cost", "Profit", "Margin %"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    records = revenue_data.get('revenue_qs', [])
    if not records:
        ws.append(["No records found", "", "", "", "", "", ""])
    else:
        for idx, record in enumerate(records, 1):
            row = [
                idx,
                f"{record.client.first_name} {record.client.last_name}",
                record.service.name,
                float(record.client_charge),
                float(record.inst_cost),
                float(record.profit_amount),
                f"{record.profit_percent:.1f}%"
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')

    # Auto Width
    for i, col_cells in enumerate(ws.columns, 1):
        column_letter = get_column_letter(i)
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            except Exception: pass
        ws.column_dimensions[column_letter].width = max_length + 2

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)
    return tmp_file.name

def revenue_excel_view(request):
    """Serve the Excel file as a download."""
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    status_filter = request.GET.get('status', 'all')

    try:
        if not start_date_str or not end_date_str:
            raise ValueError("Missing dates")
            
        # 1. Try ISO format first
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            # 2. Fallback: Try verbose format like "Jan. 1, 2025"
            clean_start = start_date_str.replace('.', '')
            clean_end = end_date_str.replace('.', '')
            start_date = datetime.strptime(clean_start, "%b %d, %Y").date()
            end_date = datetime.strptime(clean_end, "%b %d, %Y").date()
            
    except (ValueError, TypeError):
        # Fallback to current year only if parsing actually fails
        start_date = timezone.now().date().replace(month=1, day=1)
        end_date = timezone.now().date()

    file_path = generate_revenue_excel(start_date, end_date, status_filter)
    with open(file_path, 'rb') as f:
        file_data = f.read()

    response = HttpResponse(
        file_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="revenue_{start_date}_{end_date}.xlsx"'
    return response